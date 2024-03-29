from tkinter import *
import smtplib
import random
import time
import xlsxwriter
import xlrd
import openpyxl

def varify():
    opt_info = opt.get()
    if(opt_info == str(x)):
        Label(screen2,text = "").pack()
        Label(screen2,text = "Success",bg = "green").pack()
        time.sleep(1.5)
        screen2.withdraw()
        workbook = xlsxwriter.Workbook(username_info+".xlsx")
        worksheet = workbook.add_worksheet()
        worksheet.write(0,0,username_info)
        worksheet.write(0,1,password_info)
        worksheet.write(0,2,email_info)
        worksheet.write(0,3,gender_info)
        workbook.close()
        username_input.delete(0,END)
        password_input.delete(0,END)
        email_input.delete(0,END)
    else:
        Label(screen2,text = "").pack()
        Label(screen2,text = "Try again",bg = "red").pack()
        Button(screen2,text = "Resend Opt",width = "10",height = "1",command = send_mail).pack()

def send_mail():
    global screen2
    email_address = "shivamyadav6205@gmail.com"
    email_password = "kirigaya1"
    with smtplib.SMTP('smtp.gmail.com',587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        
        smtp.login(email_address,email_password)
        
        subject = "SHCS - Thank you for registering"
        global x
        x = random.randint(0000,9999)
        body = "Thank you for registering with us.\nHere is your varification code for completing further process.\n OTP -"+str(x)
        
        msg = "Subject:{}\n\n{}".format(subject,body)
        smtp.sendmail(email_address,email_info,msg)
        
    screen2 = Toplevel(screen)
    screen2.title("Varification")
    screen2.geometry("450x350")
    screen2.configure(background = '#F5FCFF')
    global opt
    opt = StringVar()
    global opt_input
    Label(screen2,text = "Enter the varification code send to your mail address").pack()
    Label(screen2,bg='#B7E9F7')
    opt_input = Entry(screen2,textvariable = opt,bg="#F5FCFF")
    opt_input.pack()
    Label(screen2,bg='#F5FCFF').pack()
    Button(screen2,text = "Varify",width = "10", height = "1",bg='#B7E9F7', command = varify).pack()
def change():
    password_info = password.get()
    a.destroy()
    v.destroy()
    w.destroy()
    Label(screen12,text = "",bg="#FFCCCB").pack()
    wbk = openpyxl.load_workbook("Shivam.xlsx")
    for wks in wbk.worksheets:
        wks.cell(row = (c+1),column = 3).value = password_info
    wbk.save("Shivam.xlsx")
    wbk.close()

def vari():
    l.destroy()
    opt_input.destroy()
    l1.destroy()
    l2.destroy()
    l3.destroy()
    global a
    global v
    global w
    opt_info = opt.get()
    if(opt_info == str(x1)):
        a = Label(screen12,text = "",bg="#FFCCCB")
        a.pack()
        v = Label(screen12,text = "Enter new password",bg = "#FFCCCB")
        v.pack()
        global password_input1
        password_input1 = Entry(screen12,width ="30", textvariable = password)
        password_input1.pack()
        w = Button(screen12,text = "Change Password",width = "15", height = "1", command = change,bg='#B7E9F7')
        w.pack()
        aq = Label(screen12,text = "success",bg="#FFCCCB")
        aq.pack()
        time.sleep(2)
        screen12.destroy()

def reset():
    global opt
    global l
    global l1
    global l2
    global l3
    opt = StringVar()
    b.destroy()
    wb = xlrd.open_workbook("Shivam.xlsx")
    sheet = wb.sheet_by_index(0)
    email_info = sheet.cell_value(c,2)
    global screen12
    screen12 = Toplevel(screen)
    screen12.title("Varification")
    screen12.geometry("500x500")
    screen12.configure(background = '#F5FCFF')
    l = Label(screen12,text = "Enter varification code sent to your registered mail address",bg="#FFCCCB")
    l.pack()
    email_address = "shivamyadav6205@gmail.com"
    email_password = "kirigaya1"
    with smtplib.SMTP('smtp.gmail.com',587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        
        smtp.login(email_address,email_password)
        
        subject = "SHCS - Thank you for registering"
        global x1
        x1 = random.randint(0000,9999)
        body = "Thank you for registering with us.\nHere is your varification code for completing further process.\n OTP -"+str(x1)
        
        msg = "Subject:{}\n\n{}".format(subject,body)
        smtp.sendmail(email_address,email_info,msg)
    l1 = Label(screen12,bg='#F5FCFF')
    l1.pack()
    global opt_input
    opt_input = Entry(screen12,textvariable = opt,bg="#F5FCFF")
    opt_input.pack()
    l2 = Label(screen12,bg='#F5FCFF')
    l2.pack()
    l3 = Button(screen12,text = "Varify",width = "15", height = "1", command = vari,bg='#B7E9F7')
    l3.pack()

def findCell(sh, searchedValue):
    global br
    br = False
    for row in range(sh.nrows):
        myCell = sh.cell(row, 0)
        if myCell.value == searchedValue:
            br = True
            c = row
            return row
    return -1

def login():
    username_info = username.get()
    password_info = password.get()
    for sh in xlrd.open_workbook("Shivam.xlsx").sheets():
        row = findCell(sh,str(username_info))
        if(br):
            break
    wb = xlrd.open_workbook("Shivam.xlsx")
    sheet = wb.sheet_by_index(0)
    if(br and username_info == sheet.cell_value(row,0) and password_info == sheet.cell_value(row,1)):
        print("Logged in")
    else:
        Label(frame4,text="",bg="#F5FCFF").pack()
        Label(frame4,text="",bg="#F5FCFF").pack()
        global b
        b = Label(frame4,text = "Username or Password incorrect",fg="red",bg="#FFCCCB")
        b.pack()
        username_input.delete(0,END)
        password_input.delete(0,END)

def login_user():
    screen.withdraw()
    global screen2
    global frame4
    width_value = screen.winfo_screenwidth()
    height_value = screen.winfo_screenheight()
    screen2 = Toplevel(screen)
    screen2.title("Varification")
    screen2.geometry("%dx%d+0+0" % (width_value,height_value))
    screen2.configure(background = '#F5FCFF')
    frame = Frame(screen2,relief=SUNKEN)
    frame.pack(side = TOP)
    frame4 = Frame(screen2,relief=SUNKEN)
    frame4.pack(side = TOP,pady=150)
    frame4.configure(background = '#B7E9F7',bd=0.5,width="800",height="400")
    lblInfo = Label(frame, font=('arial',40,'bold'),width = str(width_value),height = "1",text="Smart Health Care",bg="#B7E9F7",fg="Steel Blue", bd=10)
    lblInfo.pack()
    global username_info
    global password_info
    Label(frame4,text = "Please enter your details",bg='#B7E9F7').pack()
    Label(frame4,bg='#B7E9F7').pack()
    Label(frame4,text = "Username *",bg='#B7E9F7').pack()
    username_input = Entry(frame4,width = "30", textvariable = username)
    username_input.pack()
    Label(frame4,text = "Password *",bg='#B7E9F7').pack()
    password_input = Entry(frame4,width ="30", textvariable = password)
    password_input.pack()
    global log
    Button(frame4,text = "Login",width = "10", height = "1", command = login,bg='#B7E9F7').pack(side = LEFT)
    Button(frame4,text = "Reset Password",width = "13", height = "1", command = reset,bg='#B7E9F7').pack(side = RIGHT)

def register_user():
    global username_info
    global password_info
    global email_info
    global gender_info
    username_info = username.get()
    password_info = password.get()
    email_info = email.get()
    gender_info = gender.get()
    send_mail()

def register():
    global username
    global password
    global email
    global username_input
    global password_input
    global email_input
    global gender
    username = StringVar()
    password = StringVar()
    email = StringVar()
    gender = StringVar()
    
    
    Label(frame3,text = "Please enter your details",bg='#B7E9F7').pack()
    Label(frame3,bg='#B7E9F7').pack()
    Label(frame3,text = "Username *",bg='#B7E9F7').pack()
    username_input = Entry(frame3,width = "30", textvariable = username)
    username_input.pack()
    Label(frame3,text = "Password *",bg='#B7E9F7').pack()
    password_input = Entry(frame3,width ="30", textvariable = password)
    password_input.pack()
    Label(frame3,text = "Email *",bg='#B7E9F7').pack()
    email_input = Entry(frame3,width = "30", textvariable = email)
    email_input.pack()
    Radiobutton(frame3,justify = LEFT,text="Male   ",variable = gender,value = "male",bg='#B7E9F7').pack()
    Radiobutton(frame3,justify = LEFT,text="Female",variable = gender,value = "female",bg='#B7E9F7').pack()
    Button(frame3,text = "Register",width = "10", height = "1", command = register_user,bg='#B7E9F7').pack(side = LEFT)
    Button(frame3,text = "Login",width = "10", height = "1", command = login_user,bg='#B7E9F7').pack(side = RIGHT)

def main_screen():
    global screen
    global frame3
    screen = Tk()
    width_value = screen.winfo_screenwidth()
    height_value = screen.winfo_screenheight()
    screen.geometry("%dx%d+0+0" % (width_value,height_value))
    screen.title("Health Care")
    screen.configure(background = '#F5FCFF')
    
    frame = Frame(screen,relief=SUNKEN)
    frame.pack(side = TOP)
    frame1 = Frame(screen)
    frame1.pack(side = LEFT)
    frame3 = Frame(screen,relief=SUNKEN)
    frame3.pack(side = TOP,pady=150)
    frame3.configure(background = '#B7E9F7',bd=0.5,width="800",height="400")
    
    lblInfo1 = Label(frame, font=('arial',40,'bold'),width = str(width_value),height = "1",text="Smart Health Care",bg="#B7E9F7",fg="Steel Blue", bd=10)
    lblInfo1.pack()
    img = PhotoImage(file = "hos1.png")
    my_lable = Label(frame1,image = img)
    my_lable.pack()
    register()
    screen.mainloop()

c = 0
main_screen()